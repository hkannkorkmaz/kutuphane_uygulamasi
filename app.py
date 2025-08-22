import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import openpyxl
from collections import Counter
import matplotlib.pyplot as plt
from barcode import EAN13
from barcode.writer import ImageWriter
import os
import requests
from PIL import Image, ImageTk
from io import BytesIO
import cv2
from pyzbar import pyzbar

# ============================ Dosya & Klasörler ============================

dosya_adi = "kutuphane.xlsx"
barkod_klasor = "barkodlar"
kapak_klasor = "kapaklar"
# IP kamera geçmişi (ilk değer örnek)
gecmis_ip_urller = ["http://192.168.1.100:8080/video"]

# Şema: ["Kitap Adı","Yazar","Sayfa","ISBN","Kategori","Durum"]
BASLIK = ["Kitap Adı", "Yazar", "Sayfa", "ISBN", "Kategori", "Durum"]

if not os.path.exists(barkod_klasor):
    os.makedirs(barkod_klasor)
if not os.path.exists(kapak_klasor):
    os.makedirs(kapak_klasor)


def ensure_schema():
    """Excel yoksa oluştur; varsa 'Durum' sütunu yoksa ekle."""
    if not os.path.exists(dosya_adi):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Kütüphane"
        sheet.append(BASLIK)
        wb.save(dosya_adi)
        return

    wb = openpyxl.load_workbook(dosya_adi)
    sheet = wb.active
    # Başlık satırı yok veya eksik olabilir
    first_row = [cell.value for cell in sheet[1]]
    if not first_row or first_row[0] != "Kitap Adı":
        # Baştan yaz
        values = list(sheet.iter_rows(values_only=True))
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Kütüphane"
        sheet.append(BASLIK)
        for r in values:
            if r and any(r):
                # Eski veriler farklı düzende olabilir -> güvenli kopya
                ad = r[0] if len(r) > 0 else ""
                yazar = r[1] if len(r) > 1 else ""
                sayfa = r[2] if len(r) > 2 else ""
                isbn = r[3] if len(r) > 3 else ""
                kategori = r[4] if len(r) > 4 else ""
                durum = r[5] if len(r) > 5 and r[5] else "Okunmadı"
                sheet.append([ad, yazar, sayfa, isbn, kategori, durum])
        wb.save(dosya_adi)
        return

    # Başlık doğruysa ama 'Durum' yoksa ekle
    if "Durum" not in first_row:
        sheet.cell(row=1, column=len(first_row) + 1, value="Durum")
        for row in range(2, sheet.max_row + 1):
            if not sheet.cell(row=row, column=len(first_row) + 1).value:
                sheet.cell(row=row, column=len(first_row) + 1, value="Okunmadı")
        wb.save(dosya_adi)


ensure_schema()

# ============================ Yardımcı Fonksiyonlar ============================


import openpyxl


def listele_excel():
    """Excel’i başlık dahil döndürür ve yazar adına göre alfabetik sıralar."""
    try:
        wb = openpyxl.load_workbook("kutuphane.xlsx")
        ws = wb.active
    except Exception as e:
        print(f"Excel dosyası okunamadı: {e}")
        return []

    kitaplar = []
    for row in ws.iter_rows(values_only=True):
        kitaplar.append(list(row))

    if len(kitaplar) <= 1:
        return kitaplar  # başlık dahil bir satır varsa

    baslik = kitaplar[0]
    kitaplar_sirali = sorted(
        kitaplar[1:], key=lambda x: (x[1] or "").lower()
    )  # yazar adına göre
    return [baslik] + kitaplar_sirali


def tree_guncelle(tree_widget):
    """Treeview’i komple yeniden doldur (başlığı atlar) ve yazar adına göre alfabetik sıralar."""
    if not tree_widget:
        return
    # Mevcut tüm öğeleri temizle
    for item in tree_widget.get_children():
        tree_widget.delete(item)

    # Excel'den veriyi al
    kitaplar = listele_excel()  # sadece Excel okur

    if len(kitaplar) <= 1:
        return  # veri yoksa çık

    baslik = kitaplar[0]  # başlık
    # Yazar adına göre alfabetik sıralama
    sirali_kitaplar = sorted(kitaplar[1:], key=lambda x: (x[1] or "").lower())

    # Treeview'e ekle
    for i, kitap in enumerate(sirali_kitaplar, start=1):
        ad = kitap[0] if len(kitap) > 0 else ""
        yazar = kitap[1] if len(kitap) > 1 else ""
        sayfa = kitap[2] if len(kitap) > 2 else ""
        isbn = kitap[3] if len(kitap) > 3 else ""
        kategori = kitap[4] if len(kitap) > 4 else ""
        durum = kitap[5] if len(kitap) > 5 else "Okunmadı"
        tree_widget.insert(
            "", "end", values=(i, ad, yazar, sayfa, isbn, kategori, durum)
        )


def kapak_getir(isbn):
    """ISBN kapaklarını önbellekleyip döndürür (PhotoImage)."""
    isbn = str(isbn or "").strip()
    if not isbn:
        return None
    kapak_dosya = os.path.join(kapak_klasor, f"{isbn}.png")
    if os.path.exists(kapak_dosya):
        try:
            img = Image.open(kapak_dosya)
            img.thumbnail((120, 180))
            return ImageTk.PhotoImage(img)
        except:
            pass
    try:
        url = f"https://covers.openlibrary.org/b/isbn/{isbn}-M.jpg"
        response = requests.get(url, timeout=10)
        if response.status_code != 200:
            return None
        img = Image.open(BytesIO(response.content))
        img.thumbnail((120, 180))
        img.save(kapak_dosya)
        return ImageTk.PhotoImage(img)
    except:
        return None


def kitap_bilgisi_getir(isbn, kaynak="OpenLibrary"):
    """OpenLibrary -> GoogleBooks tekil çekim."""
    isbn = (isbn or "").replace("-", "").strip()
    if not isbn:
        return None
    if kaynak == "OpenLibrary":
        try:
            url = f"https://openlibrary.org/api/books?bibkeys=ISBN:{isbn}&format=json&jscmd=data"
            response = requests.get(url, timeout=10)
            data = response.json()
            kitap_data = data.get(f"ISBN:{isbn}")
            if not kitap_data:
                return None
            ad = kitap_data.get("title", "Bilinmiyor")
            yazar_list = kitap_data.get("authors", [])
            yazar = (
                ", ".join(
                    [y.get("name", "") for y in yazar_list if isinstance(y, dict)]
                )
                or "Bilinmiyor"
            )
            sayfa = kitap_data.get("number_of_pages", "Bilinmiyor")
            return ad, yazar, sayfa
        except:
            return None
    elif kaynak == "GoogleBooks":
        try:
            url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
            response = requests.get(url, timeout=10)
            data = response.json()
            items = data.get("items")
            if not items:
                return None
            volume_info = items[0].get("volumeInfo", {})
            ad = volume_info.get("title", "Bilinmiyor")
            yazar_list = volume_info.get("authors", [])
            yazar = ", ".join(yazar_list) if yazar_list else "Bilinmiyor"
            sayfa = volume_info.get("pageCount", "Bilinmiyor")
            return ad, yazar, sayfa
        except:
            return None
    else:
        return None


def kitap_ekle_from_api(isbn):
    """OpenLibrary -> GoogleBooks sıralı ekleme (duplikat kontrolü)."""
    mevcut_isbnler = {str(satir[3]) for satir in listele_excel()[1:] if len(satir) > 3}
    if isbn in mevcut_isbnler:
        messagebox.showinfo("Bilgi", "Bu kitap zaten ekli!")
        return False

    kitap_data = kitap_bilgisi_getir(isbn, kaynak="OpenLibrary")
    if not kitap_data:
        kitap_data = kitap_bilgisi_getir(isbn, kaynak="GoogleBooks")
    if kitap_data:
        ad, yazar, sayfa = kitap_data
        wb = openpyxl.load_workbook(dosya_adi)
        sheet = wb.active
        sheet.append([ad, yazar, sayfa, isbn, "", "Okunmadı"])
        wb.save(dosya_adi)
        tree_guncelle(tree)
        messagebox.showinfo("Başarılı", f"{ad} kitabı eklendi!")
        return True
    else:
        messagebox.showinfo("Hata", "Kitap bilgisi bulunamadı.")
        return False


def durum_degistir_by_isbn(isbn):
    """ISBN’e göre 'Durum' değerini toggle et."""
    isbn = str(isbn or "").strip()
    if not isbn:
        return
    wb = openpyxl.load_workbook(dosya_adi)
    sheet = wb.active
    # ISBN sütunu 4, Durum sütunu 6
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=4).value) == isbn:
            mevcut = sheet.cell(row=row, column=6).value
            yeni = "Okundu" if mevcut != "Okundu" else "Okunmadı"
            sheet.cell(row=row, column=6, value=yeni)
            break
    wb.save(dosya_adi)


# ============================ GUI Fonksiyonları ============================


def kitap_ekle_gui(tree_widget):
    ad = simpledialog.askstring("Kitap Adı", "Kitap Adı:")
    yazar = simpledialog.askstring("Yazar", "Yazar:")
    sayfa = simpledialog.askstring("Sayfa", "Sayfa:")
    isbn = simpledialog.askstring("ISBN", "ISBN:")
    kategori = simpledialog.askstring("Kategori", "Kategori:")
    if not ad or not yazar or not isbn:
        return
    wb = openpyxl.load_workbook(dosya_adi)
    sheet = wb.active
    sheet.append([ad, yazar, sayfa, isbn, kategori, "Okunmadı"])
    wb.save(dosya_adi)
    # Kullanıcıdan kapak resmi seçmesini iste
    from tkinter import filedialog

    kapak_dosya = filedialog.askopenfilename(
        title="Kapak Resmi Seçin (Opsiyonel)",
        filetypes=[("Resim Dosyaları", "*.jpg *.jpeg *.png *.bmp")],
    )
    if kapak_dosya:
        # Kapaklar klasörüne ISBN.png olarak kaydet
        try:
            img = Image.open(kapak_dosya)
            img.save(os.path.join(kapak_klasor, f"{isbn}.png"))
        except Exception as e:
            messagebox.showwarning("Uyarı", f"Kapak kaydedilemedi: {e}")

    tree_guncelle(tree_widget)


def kitap_sil_gui(tree_widget):
    secim = simpledialog.askinteger("Sil", "Silmek istediğiniz kitabın sıra numarası:")
    kitaplar = listele_excel()
    if not secim or secim < 1 or secim > len(kitaplar[1:]):
        return
    # Excel’de gerçek satır: başlık + secim
    wb = openpyxl.load_workbook(dosya_adi)
    sheet = wb.active
    sheet.delete_rows(secim + 1)
    wb.save(dosya_adi)
    tree_guncelle(tree_widget)
    messagebox.showinfo("Başarılı", "Seçilen kitap silindi!")


def kitap_guncelle_gui(tree_widget, kaynak="OpenLibrary"):
    try:
        wb = openpyxl.load_workbook(dosya_adi)
        sheet = wb.active
        kitaplar = list(sheet.iter_rows(values_only=True))

        for i, satir in enumerate(kitaplar[1:], start=2):  # başlık satırı = 1
            if len(satir) < 4:
                continue
            isbn = str(satir[3]).strip() if satir[3] else ""
            if not isbn:
                continue
            kategori = satir[4] if len(satir) > 4 else ""
            durum = satir[5] if len(satir) > 5 else "Okunmadı"

            kitap_data = kitap_bilgisi_getir(isbn, kaynak="OpenLibrary")
            if not kitap_data:
                kitap_data = kitap_bilgisi_getir(isbn, kaynak="GoogleBooks")
            if kitap_data:
                ad, yazar, sayfa = kitap_data
                values = [ad, yazar, sayfa, isbn, kategori, durum]
                for col, val in enumerate(values, start=1):
                    sheet.cell(row=i, column=col, value=val)

        wb.save(dosya_adi)
        tree_guncelle(tree_widget)
        messagebox.showinfo(
            "Güncelleme", "Tüm ISBN'ler internetten kontrol edildi ve güncellendi."
        )
    except PermissionError:
        messagebox.showerror(
            "Hata", "Excel dosyası açık. Lütfen kapatıp tekrar deneyin."
        )


def kitap_duzenle_gui(tree_widget):
    import openpyxl
    from tkinter import simpledialog, messagebox, filedialog

    # Düzenlenecek kitabın sıra numarasını sor
    secim = simpledialog.askinteger(
        "Düzenle", "Düzenlemek istediğiniz kitabın sıra numarası:"
    )
    kitaplar = listele_excel()
    if not secim or secim < 1 or secim > len(kitaplar[1:]):
        return

    secilen = kitaplar[secim]  # başlık satırı atlandığı için index = secim

    # Kitap bilgilerini sor, mevcut değerleri başlangıç değeri olarak koy
    ad = simpledialog.askstring("Kitap Adı", "Kitap Adı:", initialvalue=secilen[0])
    yazar = simpledialog.askstring("Yazar", "Yazar:", initialvalue=secilen[1])
    sayfa = simpledialog.askstring("Sayfa", "Sayfa:", initialvalue=secilen[2])
    isbn = simpledialog.askstring("ISBN", "ISBN:", initialvalue=secilen[3])
    kategori = simpledialog.askstring(
        "Kategori", "Kategori:", initialvalue=secilen[4] if len(secilen) > 4 else ""
    )
    durum = simpledialog.askstring(
        "Durum (Okundu/Okunmadı)",
        "Durum (Okundu/Okunmadı):",
        initialvalue=secilen[5] if len(secilen) > 5 else "Okunmadı",
    )

    if not ad or not yazar or not isbn:
        return

    # Kapak ekleme sorusu
    kapak_yolu = ""
    cevap = messagebox.askyesno("Kapak Ekle", "Kitap kapağı eklemek ister misiniz?")
    if cevap:
        kapak_yolu = filedialog.askopenfilename(
            title="Kapak Seç",
            filetypes=(
                ("PNG Dosyaları", "*.png"),
                ("JPEG Dosyaları", "*.jpg;*.jpeg"),
                ("Tüm Dosyalar", "*.*"),
            ),
        )

    # Excel dosyasını aç ve bilgileri güncelle
    wb = openpyxl.load_workbook(dosya_adi)
    sheet = wb.active
    for col, val in enumerate(
        [
            ad,
            yazar,
            sayfa,
            isbn,
            kategori,
            durum if durum in ["Okundu", "Okunmadı"] else "Okunmadı",
            kapak_yolu,  # kapak yolu ekleniyor, istersen Excel'de 7. sütun olsun
        ],
        start=1,
    ):
        sheet.cell(row=secim + 1, column=col, value=val)

    wb.save(dosya_adi)
    tree_guncelle(tree_widget)
    messagebox.showinfo("Başarılı", "Kitap güncellendi!")


def kitap_arama_gui(tree_widget):
    sorgu = simpledialog.askstring("Arama", "Kitap adı, yazar veya ISBN:")
    if not sorgu:
        return
    kitaplar = listele_excel()
    for item in tree_widget.get_children():
        tree_widget.delete(item)
    s = sorgu.lower()
    for i, satir in enumerate(kitaplar[1:], start=1):
        ad = str(satir[0]).lower()
        yazar = str(satir[1]).lower()
        isbn = str(satir[3])
        if s in ad or s in yazar or sorgu in isbn:
            durum = satir[5] if len(satir) > 5 else "Okunmadı"
            tree_widget.insert(
                "",
                "end",
                values=(i, satir[0], satir[1], satir[2], satir[3], satir[4], durum),
            )


def yazar_grafik_gui():
    kitaplar = listele_excel()
    yazarlar = [satir[1] for satir in kitaplar[1:]]
    if not yazarlar:
        messagebox.showinfo("Bilgi", "Henüz kitap eklenmedi.")
        return
    sayim = Counter(yazarlar)
    isimler = list(sayim.keys())
    adetler = list(sayim.values())
    plt.figure(figsize=(8, 5))
    plt.bar(isimler, adetler)
    plt.xlabel("Yazarlar")
    plt.ylabel("Kitap Sayısı")
    plt.title("Yazar Başına Kitap Sayısı")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.show()


def barkod_goster_gui(tree_widget):
    secim = simpledialog.askinteger(
        "Barkod", "Barkod oluşturmak için sıra numarası girin (veya iptal):"
    )
    if not secim:
        return
    kitaplar = listele_excel()
    if secim < 1 or secim > len(kitaplar[1:]):
        messagebox.showerror("Hata", "Geçersiz sıra numarası.")
        return
    satir = kitaplar[secim]  # başlık atlandı
    if len(satir) < 4 or not satir[3]:
        messagebox.showerror("Hata", "Bu satırda ISBN yok.")
        return
    isbn = str(satir[3]).replace("-", "").strip()

    # EAN13 için 12 hane (son hane checksum). ISBN-13 ise 13 hane olabilir.
    ean_girdi = None
    if len(isbn) == 12 and isbn.isdigit():
        ean_girdi = isbn
    elif len(isbn) == 13 and isbn.isdigit():
        ean_girdi = isbn[:12]
    else:
        ean_girdi = simpledialog.askstring("EAN13", "EAN13 için 12 haneli sayı girin:")
        if not ean_girdi or not ean_girdi.isdigit() or len(ean_girdi) != 12:
            messagebox.showerror("Hata", "Geçersiz EAN13 girdisi.")
            return
    try:
        ean = EAN13(ean_girdi, writer=ImageWriter())
        dosya_yolu = os.path.join(barkod_klasor, f"{isbn}.png")
        ean.save(os.path.splitext(dosya_yolu)[0])
        olasi_png = os.path.join(barkod_klasor, f"{isbn}.png")
        if os.path.exists(olasi_png):
            dosya_yolu = olasi_png
        messagebox.showinfo("Barkod", f"Barkod kaydedildi:\n{dosya_yolu}")
    except Exception as e:
        messagebox.showerror("Hata", f"Barkod oluşturulamadı.\n{e}")


def barkod_ile_cek_gui(tree_widget):
    mevcut_isbnler = {str(satir[3]) for satir in listele_excel()[1:] if len(satir) > 3}

    secim = messagebox.askquestion(
        "Seçim", "Kitabı Manuel ISBN ile eklemek istiyor musunuz?\n(Hayır: Kamera ile)"
    )
    if secim == "yes":
        isbn = simpledialog.askstring("ISBN", "ISBN Girin:")
        if not isbn:
            return
        isbn = isbn.replace("-", "").strip()
        if isbn in mevcut_isbnler:
            messagebox.showinfo("Bilgi", "Bu kitap zaten ekli!")
            return
        kitap_ekle_from_api(isbn)
        return

    # Kamera ile okuma
    kamera_secim = messagebox.askquestion(
        "Kamera Seçimi",
        "Hangi kamerayı kullanmak istersiniz?\nEvet: Dahili Kamera\nHayır: Telefon Kamerası (IP)",
    )
    if kamera_secim == "yes":
        cap = cv2.VideoCapture(0)
    else:
        ip_dialog = tk.Toplevel()
        ip_dialog.title("Telefon Kamerası URL")
        tk.Label(ip_dialog, text="IP URL girin veya seçin:").pack(padx=10, pady=5)
        url_var = tk.StringVar(value=gecmis_ip_urller[0] if gecmis_ip_urller else "")
        combobox = ttk.Combobox(
            ip_dialog, textvariable=url_var, values=gecmis_ip_urller, width=40
        )
        combobox.pack(padx=10, pady=5)

        def ip_tamam():
            ip_url = url_var.get().strip()
            if ip_url and ip_url not in gecmis_ip_urller:
                gecmis_ip_urller.insert(0, ip_url)
                if len(gecmis_ip_urller) > 5:
                    gecmis_ip_urller.pop()
            ip_dialog.destroy()
            ip_dialog.ip_url = ip_url

        tk.Button(ip_dialog, text="Tamam", command=ip_tamam).pack(pady=10)
        ip_dialog.grab_set()
        ip_dialog.wait_window()
        ip_url = getattr(ip_dialog, "ip_url", "")
        if not ip_url:
            return
        cap = cv2.VideoCapture(ip_url)

    isbn = None
    messagebox.showinfo(
        "Bilgi", "Kameraya barkodu gösterin, ESC ile iptal edebilirsiniz."
    )
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        barcodes = pyzbar.decode(frame)
        for barcode in barcodes:
            isbn_okunan = barcode.data.decode("utf-8").replace("-", "").strip()
            if isbn_okunan in mevcut_isbnler:
                messagebox.showinfo("Bilgi", "Bu kitap zaten ekli!")
                isbn_okunan = None
                break
            isbn = isbn_okunan
            cv2.rectangle(
                frame,
                (barcode.rect.left, barcode.rect.top),
                (
                    barcode.rect.left + barcode.rect.width,
                    barcode.rect.top + barcode.rect.height,
                ),
                (0, 255, 0),
                2,
            )
            cv2.putText(
                frame,
                isbn,
                (barcode.rect.left, barcode.rect.top - 10),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.5,
                (0, 255, 0),
                2,
            )
            break
        cv2.imshow("Barkod Kamerası", frame)
        if cv2.waitKey(1) & 0xFF == 27 or isbn:
            break
    cap.release()
    cv2.destroyAllWindows()
    if isbn:
        kitap_ekle_from_api(isbn)


# ============================ Tema & Görünüm ============================

THEME = {
    "bg_main": "#1b1b2f",
    "bg_left": "#111122",
    "bg_card": "#2a2a3f",
    "bg_hover": "#3c3c5c",
    "fg_text": "#f0e6d2",
    "font_normal": ("Garamond", 12, "bold"),
    "font_title": ("Garamond", 14, "bold italic"),
    "btn_font": ("Garamond", 10, "bold"),
}


def goster_kitaplar():
    """Tree veya Kart görünümü (başlık atlanır; image refs tutulur)."""
    for widget in center_frame.winfo_children():
        widget.destroy()

    kitaplar = listele_excel()
    if len(kitaplar) <= 1:
        tk.Label(
            center_frame,
            text="Kütüphane boş!",
            font=("Arial", 14, "bold"),
            fg=THEME["fg_text"],
            bg=THEME["bg_main"],
        ).pack(expand=True)
        return None

    epic_font = THEME["font_normal"]
    epic_font_title = THEME["font_title"]
    epic_bg = THEME["bg_main"]
    card_bg = THEME["bg_card"]
    hover_bg = THEME["bg_hover"]
    text_fg = THEME["fg_text"]

    if view_mode.get() == "tree":
        cols = ("No", "Kitap Adı", "Yazar", "Sayfa", "ISBN", "Kategori", "Durum")
        tree_widget = ttk.Treeview(center_frame, columns=cols, show="headings")
        style = ttk.Style()
        style.configure("Treeview.Heading", font=epic_font_title)
        style.configure("Treeview", font=epic_font, rowheight=25)
        for col in cols:
            tree_widget.heading(col, text=col)
            tree_widget.column(col, anchor="center")
        tree_widget.pack(fill="both", expand=True)
        # doldur
        for i, kitap in enumerate(kitaplar[1:], start=1):
            ad = kitap[0]
            yazar = kitap[1]
            sayfa = kitap[2]
            isbn = kitap[3]
            kategori = kitap[4] if len(kitap) > 4 else ""
            durum = kitap[5] if len(kitap) > 5 else "Okunmadı"
            tree_widget.insert(
                "", "end", values=(i, ad, yazar, sayfa, isbn, kategori, durum)
            )

        # Çift tıklama ile "Durum" toggle
        def on_tree_double_click(event):
            item = tree_widget.selection()
            if not item:
                return
            vals = tree_widget.item(item, "values")
            if len(vals) < 7:
                return
            isbn = vals[4]
            durum_degistir_by_isbn(isbn)
            tree_guncelle(tree_widget)

        tree_widget.bind("<Double-1>", on_tree_double_click)
        return tree_widget

    else:
        # Canvas + Scroll
        canvas = tk.Canvas(center_frame, bg=epic_bg, highlightthickness=0)
        scrollbar = tk.Scrollbar(center_frame, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=epic_bg)

        scroll_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Mouse wheel ile scroll
        def _on_mousewheel(event):
            if event.delta:  # Windows / MacOS
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            else:  # Linux
                if event.num == 4:
                    canvas.yview_scroll(-1, "units")
                elif event.num == 5:
                    canvas.yview_scroll(1, "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)  # Windows / Mac
        canvas.bind_all("<Button-4>", _on_mousewheel)  # Linux up
        canvas.bind_all("<Button-5>", _on_mousewheel)  # Linux down

        # Pencere genişliğine göre otomatik sütun sayısı
        frame_width = center_frame.winfo_width() or 1200
        card_width = 220
        padding = 30
        cols = max(frame_width // (card_width + padding), 1)

        image_refs = []

        for idx, kitap in enumerate(kitaplar[1:]):  # başlığı atla
            ad = kitap[0]
            yazar = kitap[1]
            sayfa = kitap[2]
            isbn = kitap[3]
            kategori = kitap[4] if len(kitap) > 4 else ""
            durum = kitap[5] if len(kitap) > 5 else "Okunmadı"

            frm = tk.Frame(scroll_frame, relief="raised", borderwidth=2, bg=card_bg)
            frm.grid(
                row=idx // cols, column=idx % cols, padx=15, pady=15, sticky="nsew"
            )

            kapak = kapak_getir(isbn)
            lbl_text = (
                f"{ad}\n{yazar}\n{sayfa} sayfa\nISBN: {isbn}\nKategori: {kategori}"
            )
            lbl = tk.Label(
                frm,
                text=lbl_text,
                justify="center",
                bg=card_bg,
                fg=text_fg,
                font=epic_font,
            )
            lbl.pack(side="top", padx=5, pady=5)

            if kapak:
                img_lbl = tk.Label(frm, image=kapak, bg=card_bg)
                img_lbl.pack(side="top", padx=5, pady=5)
                image_refs.append(kapak)

            # Durum etiketi + toggle butonu
            durum_var = tk.StringVar(value=durum)
            durum_lbl = tk.Label(
                frm,
                textvariable=durum_var,
                bg=card_bg,
                fg="#ffd166",
                font=("Garamond", 11, "bold"),
            )
            durum_lbl.pack(side="top", pady=(2, 4))

            def toggle_local(isbn_local, var_ref=durum_var):
                durum_degistir_by_isbn(isbn_local)
                # Excel’den güncel değer oku
                for sat in listele_excel()[1:]:
                    if len(sat) > 3 and str(sat[3]) == str(isbn_local):
                        var_ref.set(sat[5] if len(sat) > 5 else "Okunmadı")
                        break
                # Tree açık ise tazele
                if view_mode.get() == "tree" and isinstance(tree, ttk.Treeview):
                    tree_guncelle(tree)

            tk.Button(
                frm,
                text="Durumu Değiştir (Okundu/Okunmadı)",
                command=lambda i=isbn: toggle_local(i),
                bg=hover_bg,
                fg=text_fg,
                font=("Garamond", 10, "bold"),
                activebackground="#4a4a6a",
                activeforeground=text_fg,
            ).pack(side="top", pady=(0, 6))

            # Hover efekti
            def on_enter(e, f=frm):
                f.config(bg=hover_bg)

            def on_leave(e, f=frm):
                f.config(bg=card_bg)

            frm.bind("<Enter>", on_enter)
            frm.bind("<Leave>", on_leave)

        # Resimleri tut ki GC yemesin
        center_frame.image_refs = image_refs
        return None


# ============================ Ana GUI ============================

root = tk.Tk()
root.title("Kütüphane Uygulaması")
root.geometry("1200x700")
root.configure(bg=THEME["bg_main"])

# ============================ Tema Ayarları ============================
style = ttk.Style()
style.theme_use("clam")  # Alternatif: "alt", "default"

# Treeview için tema ayarları
style.configure(
    "Treeview",
    background="white",
    foreground="black",
    rowheight=30,
    fieldbackground="white",
)
style.configure("Treeview.Heading", font=("Arial", 12, "bold"))
style.map(
    "Treeview", background=[("selected", "#007acc")], foreground=[("selected", "white")]
)

# Combobox ve Button stilleri
style.configure("TCombobox", padding=5, relief="flat")
style.configure("TButton", font=("Arial", 11), padding=6)

# ============================ Ana Frame Yapısı ============================
left_frame = tk.Frame(root, width=220, bg=THEME["bg_left"])
left_frame.pack(side="left", fill="y")
center_frame = tk.Frame(root, bg=THEME["bg_main"])
center_frame.pack(side="right", fill="both", expand=True)

view_mode = tk.StringVar(value="card")
tree = goster_kitaplar()


def degistir_gorunumu():
    global tree
    view_mode.set("card" if view_mode.get() == "tree" else "tree")
    tree = goster_kitaplar()


# Sol menü butonları
btns = [
    ("Görünümü Değiştir", degistir_gorunumu),
    ("Kitap Ekle", lambda: kitap_ekle_gui(tree)),
    ("Kitap Sil", lambda: kitap_sil_gui(tree)),
    ("Kitap Güncelle", lambda: kitap_guncelle_gui(tree)),
    ("Kitap Düzenle", lambda: kitap_duzenle_gui(tree)),
    ("Kitap Ara", lambda: kitap_arama_gui(tree)),
    ("Yazar Grafiği", yazar_grafik_gui),
    ("Barkod Göster", lambda: barkod_goster_gui(tree)),
    ("Barkod/ISBN ile Ekle", lambda: barkod_ile_cek_gui(tree)),
]

for b, cmd in btns:
    tk.Button(
        left_frame,
        text=b,
        width=22,
        command=cmd,
        bg=THEME["bg_card"],
        fg=THEME["fg_text"],
        font=THEME["btn_font"],
        activebackground=THEME["bg_hover"],
        activeforeground=THEME["fg_text"],
    ).pack(pady=6)

# Çıkış butonu
tk.Button(
    left_frame,
    text="Çıkış",
    width=22,
    command=root.destroy,
    bg="#b33a3a",
    fg="white",
    font=("Garamond", 11, "bold"),
    activebackground="#992f2f",
    activeforeground="white",
).pack(pady=14)

root.mainloop()
